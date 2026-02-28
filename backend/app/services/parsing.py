import hashlib
import os
import re
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import get_settings
from app.services.ocr_fallback import IMAGE_EXTS, extract_ocr_text, extract_pdf_page_ocr_text
from app.services.vl_fallback import extract_image_text_with_vl, extract_pdf_page_text_with_vl, extract_pdf_text_with_vl

settings = get_settings()

_BOILERPLATE_PATTERNS = [
    re.compile(r"\bneed\s+help\??\b", flags=re.IGNORECASE),
    re.compile(r"\benquiries?\b", flags=re.IGNORECASE),
    re.compile(r"\bcomplaints?\b", flags=re.IGNORECASE),
    re.compile(r"\ball rights reserved\b", flags=re.IGNORECASE),
    re.compile(r"\bcopyright\b", flags=re.IGNORECASE),
    re.compile(r"\bpo\s*box\b", flags=re.IGNORECASE),
    re.compile(r"\bpage\s*\d+\s*(?:of|/)\s*\d+\b", flags=re.IGNORECASE),
    re.compile(r"^\s*(?:www\.|https?://)", flags=re.IGNORECASE),
]

_EN_STOPWORDS = {
    "the",
    "and",
    "of",
    "to",
    "for",
    "with",
    "in",
    "on",
    "at",
    "is",
    "are",
    "or",
    "a",
    "an",
    "by",
    "from",
    "this",
    "that",
    "your",
    "you",
    "our",
    "we",
}


def compute_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def detect_lang_simple(text: str) -> str:
    t = str(text or "")
    if not t.strip():
        return "unknown"

    zh_chars = len(re.findall(r"[\u4e00-\u9fff]", t))
    en_chars = len(re.findall(r"[A-Za-z]", t))
    tokens = [tok for tok in re.findall(r"[\u4e00-\u9fff]+|[A-Za-z]{2,}", t) if tok]
    zh_tokens = [tok for tok in tokens if re.search(r"[\u4e00-\u9fff]", tok)]
    en_tokens = [tok for tok in tokens if re.search(r"[A-Za-z]", tok)]

    total_letters = max(1, zh_chars + en_chars)
    zh_ratio = zh_chars / total_letters

    if zh_chars >= 8 and zh_ratio >= 0.18:
        return "zh"
    if len(zh_tokens) >= 4 and len(zh_tokens) >= len(en_tokens):
        return "zh"
    if len(en_tokens) >= 6 and en_chars >= 24:
        return "en"
    if en_chars > zh_chars:
        return "en"
    if zh_chars > en_chars:
        return "zh"
    return "unknown"


def _read_txt(path: str) -> str:
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ["utf-8", "utf-16", "latin-1"]:
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", errors="ignore")


def _read_pdf(path: str, max_pages: int = 80) -> str:
    pages = _read_pdf_pages(path, max_pages=max_pages)
    return "\n\n".join(pages).strip()


def _clean_text(text: str) -> str:
    raw = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    raw = re.sub(r"(?<=\w)-\n(?=\w)", "", raw)
    raw = re.sub(r"(?<=\w)\n(?=\w)", " ", raw)
    lines = [str(line or "").strip() for line in raw.split("\n")]

    rows: list[str] = []
    for line in lines:
        if not line:
            continue
        lowered = line.lower()
        if any(pattern.search(lowered) for pattern in _BOILERPLATE_PATTERNS):
            continue

        words = re.findall(r"[a-z]+", lowered)
        if len(words) >= 5 and all(w in _EN_STOPWORDS for w in words):
            continue

        if len(line) < 2:
            continue
        if re.fullmatch(r"[\W_]+", line):
            continue

        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            rows.append(normalized)
    return "\n".join(rows).strip()


def _read_pdf_pages(path: str, max_pages: int = 120) -> list[str]:
    try:
        reader = PdfReader(path)
    except Exception:
        return []

    out: list[str] = []
    limit = max(1, int(max_pages))
    for idx, page in enumerate(reader.pages):
        if idx >= limit:
            break
        text = ""
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        text = _clean_text(text)
        if not text:
            text = _clean_text(
                extract_pdf_page_ocr_text(
                    path,
                    idx,
                    dpi=int(settings.ingestion_ocr_render_dpi),
                )
            )
        if not text:
            text = _clean_text(
                extract_pdf_page_text_with_vl(
                    path,
                    idx,
                    dpi=int(settings.ingestion_ocr_render_dpi),
                )
            )
        out.append(text)
    return out


def _split_text_to_pseudo_pages(text: str, *, tokens_per_page: int = 420) -> list[str]:
    words = str(text or "").split()
    if not words:
        return []
    step = max(160, int(tokens_per_page))
    out: list[str] = []
    for i in range(0, len(words), step):
        chunk = " ".join(words[i : i + step]).strip()
        if chunk:
            out.append(chunk)
    return out


def extract_page_chunks_from_path(path: str, *, max_pages: int = 160) -> list[str]:
    ext = Path(path).suffix.lower().lstrip(".")
    if ext == "pdf":
        pages = _read_pdf_pages(path, max_pages=max_pages)
        if any(str(p or "").strip() for p in pages):
            return [p for p in pages if str(p or "").strip()]

        # Whole-document fallback when page extraction returns empty.
        ocr_text = _clean_text(extract_ocr_text(path))
        if ocr_text:
            return _split_text_to_pseudo_pages(ocr_text, tokens_per_page=420)
        vl_text = _clean_text(
            extract_pdf_text_with_vl(
                path,
                max_pages=min(max(1, int(max_pages)), int(settings.ingestion_ocr_pdf_max_pages)),
                dpi=int(settings.ingestion_ocr_render_dpi),
            )
        )
        if vl_text:
            return _split_text_to_pseudo_pages(vl_text, tokens_per_page=420)
        return []

    if ext in {"txt", "md"}:
        return _split_text_to_pseudo_pages(_read_txt(path), tokens_per_page=420)
    if ext == "docx":
        return _split_text_to_pseudo_pages(_read_docx(path), tokens_per_page=420)
    if ext == "xlsx":
        return _split_text_to_pseudo_pages(_read_xlsx(path), tokens_per_page=420)
    if ext in IMAGE_EXTS:
        text = _clean_text(extract_ocr_text(path))
        if not text:
            text = _clean_text(extract_image_text_with_vl(path))
        return [text] if text else []
    raise ValueError(f"unsupported_extension:{ext}")


def _read_docx(path: str) -> str:
    doc = DocxDocument(path)
    return "\n".join(p.text for p in doc.paragraphs if str(p.text).strip())


def _read_xlsx(path: str, max_rows: int = 2000) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    row_count = 0
    for ws in wb.worksheets:
        parts.append(f"## Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            if row_count >= max_rows:
                break
            vals = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if vals:
                parts.append(" | ".join(vals))
                row_count += 1
    return "\n".join(parts)


def extract_text_from_path(path: str) -> str:
    ext = Path(path).suffix.lower().lstrip(".")
    if ext in {"txt", "md"}:
        return _read_txt(path)
    if ext == "pdf":
        pages = extract_page_chunks_from_path(path, max_pages=120)
        text = "\n\n".join(f"[Page {idx + 1}]\n{page}" for idx, page in enumerate(pages) if str(page or "").strip())
        if text.strip():
            return text
        return _clean_text(extract_ocr_text(path))
    if ext == "docx":
        return _read_docx(path)
    if ext == "xlsx":
        return _read_xlsx(path)
    if ext in IMAGE_EXTS:
        text = _clean_text(extract_ocr_text(path))
        if text:
            return text
        return _clean_text(extract_image_text_with_vl(path))
    raise ValueError(f"unsupported_extension:{ext}")


def chunk_text(text: str, target_tokens: int = 320, overlap_tokens: int = 48) -> list[str]:
    tokens = str(text or "").split()
    if not tokens:
        return []

    if target_tokens < 40:
        target_tokens = 40
    if overlap_tokens < 0:
        overlap_tokens = 0
    if overlap_tokens >= target_tokens:
        overlap_tokens = target_tokens // 4

    chunks: list[str] = []
    step = max(1, target_tokens - overlap_tokens)
    i = 0
    while i < len(tokens):
        part = tokens[i : i + target_tokens]
        if not part:
            break
        chunks.append(" ".join(part).strip())
        i += step
    return chunks


def file_meta(path: str) -> tuple[str, str, int]:
    p = Path(path)
    return (p.name, p.suffix.lower().lstrip("."), os.path.getsize(path))


def build_bilingual_title(file_name: str) -> tuple[str, str]:
    base = os.path.splitext(file_name)[0].strip()
    if not base:
        base = "Untitled"
    return (base, base)
